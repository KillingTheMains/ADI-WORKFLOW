"""
show_crew.py — routes for assigning crew members to a specific show.

URL prefix: /shows/<show_id>/crew
"""
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from extensions import db
from models import Show, CrewMember, ShowCrewAssignment, Company, Position, \
    ScheduleActivity, CrewRow, ShowOpenSlot
from datetime import date as date_cls
from crew_ordering import crew_order_by, crew_sort_key

show_crew_bp = Blueprint("show_crew", __name__)


# ── Show crew roster page ─────────────────────────────────────────────────────
#
# (Removed a dead `_get_show_or_404` helper that raised a bare Exception
# instead of a real 404. Every route uses `Show.query.get_or_404` directly.)

@show_crew_bp.route("/<int:show_id>/crew")
def show_crew(show_id):
    show = Show.query.get_or_404(show_id)

    # All active crew, grouped by company
    all_crew = (
        db.session.query(CrewMember)
        .filter_by(active=True)
        .outerjoin(Position, CrewMember.position_id == Position.id)
        .order_by(*crew_order_by())
        .all()
    )

    # Set of crew_member_ids already assigned to this show
    assigned_ids = {a.crew_member_id for a in show.crew_assignments}

    # Group crew by company for display
    companies = {}
    for cm in all_crew:
        co_name = cm.company.name if cm.company else "No Company"
        co_id   = cm.company_id or 0
        if co_id not in companies:
            companies[co_id] = {"name": co_name, "crew": []}
        companies[co_id]["crew"].append(cm)

    # Sort companies alphabetically
    sorted_companies = sorted(companies.values(), key=lambda c: c["name"])

    # ── Phase A: booked crew + open slots, grouped by booking_task ──────────
    assignments = (ShowCrewAssignment.query
                   .filter_by(show_id=show_id)
                   .all())
    open_slots  = (ShowOpenSlot.query
                   .filter_by(show_id=show_id)
                   .all())

    # ── #29 Phase 2: roster grouped by COMPANY; open/TBD slots collect in
    # their own group at the end (Option A). Within a company, manual drag
    # (assignment.sort_order) wins, else the canonical Crew Database order.
    roster_map = {}
    for a in assignments:
        cm = a.crew_member
        co_id = (cm.company_id if cm else None) or 0
        co_name = cm.company.name if cm and cm.company else "No Company"
        g = roster_map.setdefault(co_id, {"name": co_name, "assignments": [],
                                          "slots": [], "rows": []})
        g["assignments"].append(a)
        g["rows"].append({
            "kind": "a", "obj": a,
            "_sort": ((a.sort_order if a.sort_order is not None else 10**9,)
                      + (crew_sort_key(cm) if cm else (10**9,))),
        })
    for g in roster_map.values():
        g["rows"].sort(key=lambda r: r["_sort"])
    roster_groups = [g for _, g in sorted(roster_map.items(),
                                          key=lambda kv: kv[1]["name"].lower())]

    # Open/TBD slots have no company → one group at the end of the roster.
    if open_slots:
        tbd = {"name": "Open Positions (TBD)", "assignments": [],
               "slots": [], "rows": []}
        for s in sorted(open_slots, key=lambda s: (
                s.sort_order if s.sort_order is not None else 10**9, s.id)):
            tbd["slots"].append(s)
            tbd["rows"].append({"kind": "s", "obj": s, "_sort": ()})
        roster_groups.append(tbd)

    # Position list for the "+ TBD slot" picker
    all_positions = Position.query.order_by(Position.department, Position.title).all()

    return render_template(
        "shows/show_crew.html",
        show=show,
        sorted_companies=sorted_companies,
        assigned_ids=assigned_ids,
        roster_groups=roster_groups,
        all_positions=all_positions,
        all_companies=Company.query.order_by(Company.name).all(),
    )


# ── Assign / unassign a single crew member ────────────────────────────────────

def _autofill_travel_window(assignment, show):
    """#31 — fill Travel In/Out from the show's designated travel window, but
    ONLY when blank. Additive: never overwrites a date already set."""
    if show.travel_window_start and not assignment.travel_in_date:
        assignment.travel_in_date = show.travel_window_start
    if show.travel_window_end and not assignment.travel_out_date:
        assignment.travel_out_date = show.travel_window_end
    return assignment


@show_crew_bp.route("/<int:show_id>/crew/assign", methods=["POST"])
def assign_crew(show_id):
    show = Show.query.get_or_404(show_id)
    crew_member_id = int(request.form["crew_member_id"])
    action = request.form.get("action", "assign")  # "assign" or "unassign"

    if action == "assign":
        existing = ShowCrewAssignment.query.filter_by(
            show_id=show_id, crew_member_id=crew_member_id).first()
        if not existing:
            a = ShowCrewAssignment(show_id=show_id, crew_member_id=crew_member_id)
            _autofill_travel_window(a, show)
            db.session.add(a)
            db.session.commit()
    else:
        ShowCrewAssignment.query.filter_by(
            show_id=show_id, crew_member_id=crew_member_id).delete()
        db.session.commit()

    return redirect(url_for("show_crew.show_crew", show_id=show_id))


# ── Assign / unassign an entire company in one click ─────────────────────────

@show_crew_bp.route("/<int:show_id>/crew/assign-company", methods=["POST"])
def assign_company(show_id):
    show = Show.query.get_or_404(show_id)
    company_id = int(request.form["company_id"])
    action = request.form.get("action", "assign")

    company_crew = CrewMember.query.filter_by(
        company_id=company_id, active=True).all()

    if action == "assign":
        for cm in company_crew:
            exists = ShowCrewAssignment.query.filter_by(
                show_id=show_id, crew_member_id=cm.id).first()
            if not exists:
                a = ShowCrewAssignment(show_id=show_id, crew_member_id=cm.id)
                _autofill_travel_window(a, show)
                db.session.add(a)
        db.session.commit()
        co = Company.query.get(company_id)
        flash(f"Added all {co.name} crew to {show.name}.", "success")
    else:
        crew_ids = [cm.id for cm in company_crew]
        ShowCrewAssignment.query.filter(
            ShowCrewAssignment.show_id == show_id,
            ShowCrewAssignment.crew_member_id.in_(crew_ids)
        ).delete(synchronize_session=False)
        db.session.commit()
        co = Company.query.get(company_id)
        flash(f"Removed all {co.name} crew from {show.name}.", "info")

    return redirect(url_for("show_crew.show_crew", show_id=show_id))


# ── AJAX: Add all assigned crew from a company to a specific activity ─────────

@show_crew_bp.route(
    "/<int:show_id>/schedule/<int:day_id>/activities/<int:act_id>/add-company-crew",
    methods=["POST"])
def add_company_crew_to_activity(show_id, day_id, act_id):
    """
    Bulk-add every show-assigned crew member from a company to an activity.
    Called via AJAX from the day editor.
    """
    activity   = ScheduleActivity.query.get_or_404(act_id)
    company_id = int(request.json.get("company_id", 0))
    hours      = request.json.get("hours")          # float or None
    if hours is not None:
        try:
            hours = float(hours)
        except (TypeError, ValueError):
            hours = None

    # Crew assigned to this show AND belonging to this company
    assigned = (
        db.session.query(CrewMember)
        .join(ShowCrewAssignment,
              ShowCrewAssignment.crew_member_id == CrewMember.id)
        .filter(
            ShowCrewAssignment.show_id == show_id,
            CrewMember.company_id == company_id,
            CrewMember.active == True,
        )
        .outerjoin(Position, CrewMember.position_id == Position.id)
        .order_by(*crew_order_by())
        .all()
    )

    # Current max sort_order in this activity
    existing_max = db.session.query(
        db.func.max(CrewRow.sort_order)
    ).filter_by(activity_id=act_id).scalar() or 0

    added = []
    for i, cm in enumerate(assigned):
        # Skip if already on this activity
        already = CrewRow.query.filter_by(
            activity_id=act_id, crew_member_id=cm.id).first()
        if already:
            continue
        row = CrewRow(
            activity_id=act_id,
            crew_member_id=cm.id,
            position=cm.position.title if cm.position else "",
            sort_order=existing_max + i + 1,
            hours=hours,
        )
        db.session.add(row)
        added.append({
            "id": None,  # filled after commit
            "crew_member_id": cm.id,
            "name": cm.full_name,
            "position": cm.position.title if cm.position else "",
        })

    db.session.commit()

    # Fill in the real IDs
    for item in added:
        row = CrewRow.query.filter_by(
            activity_id=act_id,
            crew_member_id=item["crew_member_id"]
        ).first()
        if row:
            item["id"] = row.id

    return jsonify({"added": len(added), "rows": added})


# ── Crew contact sheet ────────────────────────────────────────────────────────

@show_crew_bp.route("/<int:show_id>/crew/contact-sheet")
def contact_sheet(show_id):
    """Printable contact sheet for all crew assigned to this show."""
    show = Show.query.get_or_404(show_id)

    # Pull assigned crew with their full relationships loaded
    assignments = (
        db.session.query(ShowCrewAssignment)
        .join(CrewMember, ShowCrewAssignment.crew_member_id == CrewMember.id)
        .outerjoin(Position, CrewMember.position_id == Position.id)
        .filter(ShowCrewAssignment.show_id == show_id, CrewMember.active == True)
        .order_by(*crew_order_by())
        .all()
    )

    # Group by company
    companies = {}
    for a in assignments:
        cm = a.crew_member
        co_name = cm.company.name if cm.company else "No Company"
        co_id   = cm.company_id or 0
        if co_id not in companies:
            companies[co_id] = {"name": co_name, "crew": []}
        companies[co_id]["crew"].append(cm)

    sorted_companies = sorted(companies.values(), key=lambda c: c["name"])

    return render_template(
        "shows/crew_contact_sheet.html",
        show=show,
        sorted_companies=sorted_companies,
        total=len(assignments),
    )


# ── Show hours report ─────────────────────────────────────────────────────────

@show_crew_bp.route("/<int:show_id>/crew/hours")
def hours_report(show_id):
    """Per-crew-member hours breakdown across all show days."""
    show = Show.query.get_or_404(show_id)

    # Build a lookup: crew_member_id → {member, days: {day_id: hours}, total}
    crew_data = {}   # keyed by crew_member_id (for named crew)
    tbd_data  = []   # list of {name, day_id, hours, position, activity} for unnamed rows

    show_days = show.days  # already ordered by date

    for day in show_days:
        for act in day.activities:
            for row in act.crew_rows:
                if row.is_group_header:
                    continue
                qty     = row.qty or 1
                hrs     = (row.hours or 0) * qty
                actual  = (row.actual_hours or 0) * qty
                if row.crew_member_id:
                    if row.crew_member_id not in crew_data:
                        cm = row.crew_member
                        crew_data[row.crew_member_id] = {
                            "member":   cm,
                            "position": row.position or (cm.position.title if cm.position else ""),
                            "dept":     (cm.position.department if cm.position else "") or "",
                            "company":  cm.company.name if cm.company else "",
                            "type":     row.crew_type or "",
                            "days":     {},
                            "total":    0.0,
                            "total_actual": 0.0,
                            "actual_recorded": False,
                        }
                    entry = crew_data[row.crew_member_id]
                    entry["days"][day.id] = entry["days"].get(day.id, 0.0) + hrs
                    entry["total"] += hrs
                    entry["total_actual"] += actual
                    if row.actual_hours is not None:
                        entry["actual_recorded"] = True
                else:
                    # TBD / local unnamed row — track separately
                    tbd_data.append({
                        "name":     row.display_name,
                        "position": row.position or "",
                        "type":     row.crew_type or "",
                        "day_id":   day.id,
                        "hours":    hrs,
                        "actual":   actual if row.actual_hours is not None else None,
                        "activity": act.description,
                    })

    # Sort named crew: by company, then the canonical Crew Database order (#29)
    sorted_crew = sorted(
        crew_data.values(),
        key=lambda x: (x["company"], crew_sort_key(x["member"]))
    )

    # Day totals (sum of all named crew hours per day)
    day_totals = {}
    for entry in sorted_crew:
        for day_id, hrs in entry["days"].items():
            day_totals[day_id] = day_totals.get(day_id, 0.0) + hrs

    grand_total        = sum(e["total"] for e in sorted_crew)
    grand_total_actual = sum(e["total_actual"] for e in sorted_crew)
    any_actual_recorded = any(e.get("actual_recorded") for e in sorted_crew)

    return render_template(
        "shows/hours_report.html",
        show=show,
        show_days=show_days,
        sorted_crew=sorted_crew,
        tbd_data=tbd_data,
        day_totals=day_totals,
        grand_total=grand_total,
        grand_total_actual=grand_total_actual,
        any_actual_recorded=any_actual_recorded,
    )



# ── Phase A: edit booking info on an existing assignment ─────────────────────

def _parse_date(s):
    s = (s or "").strip()
    if not s:
        return None
    try:
        return date_cls.fromisoformat(s)
    except ValueError:
        return None


def _set_if_present(obj, attr, form, key, transform=None):
    """Only update attr if the form key is in the request — lets booking
    page and travel page each post their own slice of fields without
    blanking the other's data."""
    if key not in form:
        return
    raw = form.get(key) or ""
    val = transform(raw) if transform else (raw.strip() or None)
    setattr(obj, attr, val)


def _to_float(s):
    s = (s or "").strip().replace("$", "").replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


@show_crew_bp.route("/<int:show_id>/crew/assignment/<int:aid>/edit", methods=["POST"])
def edit_assignment(show_id, aid):
    a = ShowCrewAssignment.query.get_or_404(aid)
    if a.show_id != show_id:
        flash("Assignment doesn't belong to this show.", "danger")
        return redirect(url_for("show_crew.show_crew", show_id=show_id))
    f = request.form
    next_url = (f.get("next") or "").strip()

    # Booking fields (Booking Sheet form)
    _set_if_present(a, "booking_task",   f, "booking_task")
    _set_if_present(a, "role_override",  f, "role_override")
    _set_if_present(a, "travel_in_date", f, "travel_in_date", _parse_date)
    _set_if_present(a, "start_date",     f, "start_date",     _parse_date)
    _set_if_present(a, "end_date",       f, "end_date",       _parse_date)
    _set_if_present(a, "travel_out_date",f, "travel_out_date",_parse_date)
    # Travel fields (Travel page form)
    _set_if_present(a, "hotel_name",         f, "hotel_name")
    _set_if_present(a, "hotel_check_in",     f, "hotel_check_in",  _parse_date)
    _set_if_present(a, "hotel_check_out",    f, "hotel_check_out", _parse_date)
    _set_if_present(a, "hotel_confirmation", f, "hotel_confirmation")
    _set_if_present(a, "hotel_cost",         f, "hotel_cost",      _to_float)
    _set_if_present(a, "arrival_flight",     f, "arrival_flight")
    _set_if_present(a, "arrival_time",       f, "arrival_time")
    _set_if_present(a, "departure_flight",   f, "departure_flight")
    _set_if_present(a, "departure_time",     f, "departure_time")
    _set_if_present(a, "itinerary_link",     f, "itinerary_link")

    db.session.commit()
    flash(f"Saved {a.crew_member.full_name}.", "success")
    # Respect a posted `next=` (Travel page submits it) so we land back
    # where the user came from.
    if next_url and next_url.startswith("/"):
        return redirect(next_url)
    return redirect(url_for("show_crew.show_crew", show_id=show_id))


# ── Phase A: TBD / open-slot CRUD ────────────────────────────────────────────

@show_crew_bp.route("/<int:show_id>/crew/slot/add", methods=["POST"])
def add_slot(show_id):
    show = Show.query.get_or_404(show_id)
    f = request.form
    pos_id = (f.get("position_id") or "").strip()
    slot = ShowOpenSlot(
        show_id          = show_id,
        position_id      = int(pos_id) if pos_id.isdigit() else None,
        placeholder_label= (f.get("placeholder_label") or "").strip() or None,
        booking_task     = (f.get("booking_task") or "").strip() or None,
        travel_in_date   = _parse_date(f.get("travel_in_date")),
        start_date       = _parse_date(f.get("start_date")),
        end_date         = _parse_date(f.get("end_date")),
        travel_out_date  = _parse_date(f.get("travel_out_date")),
        notes            = (f.get("notes") or "").strip() or None,
    )
    if not slot.position_id and not slot.placeholder_label:
        flash("Pick a Position OR enter a label for the slot.", "danger")
        return redirect(url_for("show_crew.show_crew", show_id=show_id))
    db.session.add(slot)
    db.session.commit()
    flash(f"Added TBD slot: {slot.display_title}.", "success")
    return redirect(url_for("show_crew.show_crew", show_id=show_id))


@show_crew_bp.route("/<int:show_id>/crew/slot/<int:sid>/edit", methods=["POST"])
def edit_slot(show_id, sid):
    slot = ShowOpenSlot.query.get_or_404(sid)
    if slot.show_id != show_id:
        flash("Slot doesn't belong to this show.", "danger")
        return redirect(url_for("show_crew.show_crew", show_id=show_id))
    f = request.form
    pos_id = (f.get("position_id") or "").strip()
    slot.position_id      = int(pos_id) if pos_id.isdigit() else None
    slot.placeholder_label= (f.get("placeholder_label") or "").strip() or None
    slot.booking_task     = (f.get("booking_task") or "").strip() or None
    slot.travel_in_date   = _parse_date(f.get("travel_in_date"))
    slot.start_date       = _parse_date(f.get("start_date"))
    slot.end_date         = _parse_date(f.get("end_date"))
    slot.travel_out_date  = _parse_date(f.get("travel_out_date"))
    slot.notes            = (f.get("notes") or "").strip() or None
    db.session.commit()
    flash("Slot updated.", "success")
    return redirect(url_for("show_crew.show_crew", show_id=show_id))


@show_crew_bp.route("/<int:show_id>/crew/slot/<int:sid>/delete", methods=["POST"])
def delete_slot(show_id, sid):
    slot = ShowOpenSlot.query.get_or_404(sid)
    if slot.show_id != show_id:
        flash("Slot doesn't belong to this show.", "danger")
        return redirect(url_for("show_crew.show_crew", show_id=show_id))
    db.session.delete(slot)
    db.session.commit()
    flash("Slot removed.", "success")
    return redirect(url_for("show_crew.show_crew", show_id=show_id))


@show_crew_bp.route("/<int:show_id>/crew/slot/<int:sid>/fill", methods=["POST"])
def fill_slot(show_id, sid):
    """Convert a TBD slot into a real ShowCrewAssignment, carrying the
    slot's booking_task and date window over to the new assignment."""
    slot = ShowOpenSlot.query.get_or_404(sid)
    if slot.show_id != show_id:
        flash("Slot doesn't belong to this show.", "danger")
        return redirect(url_for("show_crew.show_crew", show_id=show_id))
    cm_id = (request.form.get("crew_member_id") or "").strip()
    if not cm_id.isdigit():
        flash("Pick a crew member to fill this slot.", "danger")
        return redirect(url_for("show_crew.show_crew", show_id=show_id))

    # Don't double-assign the same person to the same show
    existing = ShowCrewAssignment.query.filter_by(
        show_id=show_id, crew_member_id=int(cm_id)).first()
    if existing:
        # Just merge the slot's booking info into the existing assignment
        existing.booking_task    = existing.booking_task    or slot.booking_task
        existing.travel_in_date  = existing.travel_in_date  or slot.travel_in_date
        existing.start_date      = existing.start_date      or slot.start_date
        existing.end_date        = existing.end_date        or slot.end_date
        existing.travel_out_date = existing.travel_out_date or slot.travel_out_date
    else:
        db.session.add(ShowCrewAssignment(
            show_id          = show_id,
            crew_member_id   = int(cm_id),
            booking_task     = slot.booking_task,
            travel_in_date   = slot.travel_in_date,
            start_date       = slot.start_date,
            end_date         = slot.end_date,
            travel_out_date  = slot.travel_out_date,
        ))
    db.session.delete(slot)
    db.session.commit()
    flash("Slot filled.", "success")
    return redirect(url_for("show_crew.show_crew", show_id=show_id))



# ── Phase B: Travel page (per-crew hotel + flight detail) ────────────────────

def _travel_assignments_sorted(show_id, sort_by="check_in"):
    """Return this show's assignments in the given sort order."""
    items = ShowCrewAssignment.query.filter_by(show_id=show_id).all()

    def _name(a):
        return (a.crew_member.last_name or "").lower() if a.crew_member else ""
    def _company(a):
        cm = a.crew_member
        return (cm.company.name or "").lower() if cm and cm.company else "zzz"
    def _position(a):
        cm = a.crew_member
        return (cm.position.title or "").lower() if cm and cm.position else "zzz"
    def _order(a):   # canonical Crew Database order (#29)
        cm = a.crew_member
        return cm.sort_order if (cm and cm.sort_order is not None) else 10**9

    if sort_by == "company":
        items.sort(key=lambda a: (_company(a), _order(a), _name(a)))
    elif sort_by == "name":
        items.sort(key=_name)
    elif sort_by == "position":
        items.sort(key=lambda a: (_position(a), _name(a)))
    else:   # check_in — default (None → bottom, then by name)
        # Check-in now mirrors the shared Travel In date.
        items.sort(key=lambda a: (a.travel_in_date or date_cls.max, _name(a)))
    return items


def _company_name(a):
    """Display name of an assignment's company ('No Company' when unset)."""
    cm = a.crew_member
    return cm.company.name if cm and cm.company else "No Company"


def _company_counts(assignments):
    """{company_name: traveler_count} for the on-screen company banners."""
    counts = {}
    for a in assignments:
        name = _company_name(a)
        counts[name] = counts.get(name, 0) + 1
    return counts


@show_crew_bp.route("/<int:show_id>/crew/travel")
def travel(show_id):
    show = Show.query.get_or_404(show_id)
    sort_by = (request.args.get("sort") or "check_in").strip().lower()
    if sort_by not in ("check_in", "name", "company", "position"):
        sort_by = "check_in"
    assignments = _travel_assignments_sorted(show_id, sort_by)
    grand_total  = sum((a.hotel_cost or 0) for a in assignments)
    grand_nights = sum((a.stay_nights or 0) for a in assignments)
    return render_template("shows/show_crew_travel.html",
                           show=show,
                           assignments=assignments,
                           grand_total=grand_total,
                           grand_nights=grand_nights,
                           company_counts=_company_counts(assignments),
                           sort_by=sort_by)


@show_crew_bp.route("/<int:show_id>/crew/travel/bulk-dates", methods=["POST"])
def travel_bulk_dates(show_id):
    """Set Travel-In / Show-in (start) / Show-out (end) / Travel-Out on many
    crew at once. Larry's request: 'Select specific crew members or ALL, and
    set travel in, show, travel out dates for several or all crew at once.'

    Only the date fields that are actually filled in the toolbar are applied;
    blank fields leave each row's existing value alone (so you can push just a
    travel-in date to a group without wiping their return dates). Every write
    goes through the normal SQLAlchemy path, so the audit log captures it and
    it's undoable from Recent Activity."""
    Show.query.get_or_404(show_id)
    f = request.form

    # Which rows? Support checkbox lists posted as assignment_ids.
    raw_ids = f.getlist("assignment_ids")
    ids = set()
    for r in raw_ids:
        try:
            ids.add(int(r))
        except (TypeError, ValueError):
            continue

    # Map of model-attr → parsed date, keeping only non-empty values.
    field_map = {
        "travel_in_date":  _parse_date(f.get("travel_in_date")),
        "start_date":      _parse_date(f.get("start_date")),
        "end_date":        _parse_date(f.get("end_date")),
        "travel_out_date": _parse_date(f.get("travel_out_date")),
    }
    updates = {k: v for k, v in field_map.items() if v is not None}

    sort_by = (f.get("sort") or "check_in").strip().lower()
    if sort_by not in ("check_in", "name", "company", "position"):
        sort_by = "check_in"
    back = url_for("show_crew.travel", show_id=show_id, sort=sort_by)

    if not ids:
        flash("No crew selected — pick at least one row, then apply.", "warning")
        return redirect(back)
    if not updates:
        flash("No dates entered — fill at least one date field, then apply.", "warning")
        return redirect(back)

    # Only touch assignments that belong to THIS show (defensive).
    rows = (ShowCrewAssignment.query
            .filter(ShowCrewAssignment.show_id == show_id,
                    ShowCrewAssignment.id.in_(ids))
            .all())
    n = 0
    for a in rows:
        for attr, val in updates.items():
            setattr(a, attr, val)
        n += 1
    db.session.commit()

    which = ", ".join(k.replace("_date", "").replace("_", " ") for k in updates)
    flash(f"Updated {which} on {n} crew member{'' if n == 1 else 's'}. "
          f"Undo from Recent Activity if needed.", "success")
    return redirect(back)



# ── Drag-to-reorder on the Booking Sheet ─────────────────────────────────────

@show_crew_bp.route("/<int:show_id>/crew/reorder", methods=["POST"])
def reorder(show_id):
    """
    Bulk-update sort_order from a drag-and-drop reorder within a booking-task
    card. Accepts JSON:
       { "items": [{"type": "a"|"s", "id": <int>}, ...] }
    Assigns sort_order = idx * 10 to each item in the given order. Rows
    from a different show are ignored defensively.
    """
    from flask import jsonify
    data = request.get_json(silent=True) or {}
    items = data.get("items") or []
    if not isinstance(items, list):
        return jsonify(ok=False, error="items must be a list"), 400
    n = 0
    for idx, it in enumerate(items):
        if not isinstance(it, dict):
            continue
        kind = it.get("type")
        try:
            rid = int(it.get("id"))
        except (TypeError, ValueError):
            continue
        if kind == "a":
            obj = ShowCrewAssignment.query.get(rid)
        elif kind == "s":
            obj = ShowOpenSlot.query.get(rid)
        else:
            continue
        if obj and obj.show_id == show_id:
            obj.sort_order = idx * 10
            n += 1
    db.session.commit()
    return jsonify(ok=True, count=n)



# ── Contact sheet + Travel exports (XLSX / PDF) ──────────────────────────────
import io as _io
from flask import send_file


def _xlsx_response(wb, filename):
    """Serialize an openpyxl Workbook and stream it as a download."""
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _slugify(s):
    return "".join(c if c.isalnum() else "_" for c in (s or "")).strip("_")


@show_crew_bp.route("/<int:show_id>/crew/contact-sheet.xlsx")
def contact_sheet_xlsx(show_id):
    """Same data as the on-screen contact sheet, delivered as XLSX."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    show = Show.query.get_or_404(show_id)
    assignments = (
        db.session.query(ShowCrewAssignment)
        .join(CrewMember, ShowCrewAssignment.crew_member_id == CrewMember.id)
        .outerjoin(Position, CrewMember.position_id == Position.id)
        .filter(ShowCrewAssignment.show_id == show_id, CrewMember.active == True)
        .order_by(*crew_order_by())
        .all()
    )
    # Group by company (same as the HTML view)
    companies = {}
    for a in assignments:
        cm = a.crew_member
        co_name = cm.company.name if cm.company else "No Company"
        co_id   = cm.company_id or 0
        companies.setdefault(co_id, {"name": co_name, "crew": []})["crew"].append(cm)
    sorted_companies = sorted(companies.values(), key=lambda c: c["name"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contact Sheet"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A1A1A")
    company_fill = PatternFill("solid", fgColor="F5F5F5")

    # Title row
    ws.append([f"{show.code or ''}   {show.name}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Crew Contact Sheet   ·   "
               f"{show.venue.name + ' — ' + show.venue.city if show.venue else ''}"])
    ws.append([])

    headers = ["Name", "Position", "Department", "Phone", "Email"]
    for co in sorted_companies:
        # Company banner
        ws.append([f"{co['name']}   —   {len(co['crew'])} member(s)"])
        row = ws.max_row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = header_fill

        # Header row
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=c).font = Font(bold=True, size=9)
            ws.cell(row=ws.max_row, column=c).fill = company_fill

        # Data
        for cm in co["crew"]:
            ws.append([
                cm.full_name,
                cm.position.title if cm.position else "",
                cm.position.department if cm.position else "",
                cm.phone or "",
                cm.email or "",
            ])
        ws.append([])

    # Column widths
    widths = [22, 18, 14, 15, 34]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    fname = f"{_slugify(show.code or show.name)}_contact_sheet.xlsx"
    return _xlsx_response(wb, fname)


@show_crew_bp.route("/<int:show_id>/crew/travel.xlsx")
def travel_xlsx(show_id):
    """Travel table as XLSX, respecting the ?sort= query param."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    show = Show.query.get_or_404(show_id)
    sort_by = (request.args.get("sort") or "check_in").strip().lower()
    if sort_by not in ("check_in", "name", "company", "position"):
        sort_by = "check_in"
    assignments = _travel_assignments_sorted(show_id, sort_by)
    company_counts = _company_counts(assignments)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Travel"

    ws.append([f"{show.code or ''}   {show.name}   —   Travel"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Sorted by: {sort_by.replace('_', ' ')}"])
    ws.append([])

    headers = ["Name", "Company", "Position", "Booking Task",
               "Hotel", "Check In", "Check Out", "Nights",
               "Conf #", "Cost",
               "Arr Flight #", "Arr Time", "Dep Flight #", "Dep Time",
               "Itinerary"]

    def _write_header_row():
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="1A1A1A")

    company_banner = (sort_by == "company")
    company_fill = PatternFill("solid", fgColor="F5F5F5")

    if not company_banner:
        _write_header_row()

    cur_company = None
    for a in assignments:
        cm = a.crew_member
        # When sorted by company, emit a section banner + header row on change,
        # mirroring the Crew Contact Sheet layout.
        if company_banner:
            co_name = _company_name(a)
            if co_name != cur_company:
                cur_company = co_name
                n = company_counts.get(co_name, 0)
                ws.append([f"{co_name}   —   {n} traveler{'' if n == 1 else 's'}"])
                brow = ws.max_row
                ws.merge_cells(start_row=brow, start_column=1,
                               end_row=brow, end_column=len(headers))
                bcell = ws.cell(row=brow, column=1)
                bcell.font = Font(bold=True, color="FFFFFF")
                bcell.fill = PatternFill("solid", fgColor="1A1A1A")
                _write_header_row()
        ws.append([
            cm.full_name if cm else "",
            cm.company.name if cm and cm.company else "",
            cm.position.title if cm and cm.position else "",
            a.booking_task or "",
            a.hotel_name or "",
            a.travel_in_date.isoformat() if a.travel_in_date else "",
            a.travel_out_date.isoformat() if a.travel_out_date else "",
            a.stay_nights if a.stay_nights is not None else "",
            a.hotel_confirmation or "",
            a.hotel_cost if a.hotel_cost is not None else "",
            a.arrival_flight or "",
            a.arrival_time or "",
            a.departure_flight or "",
            a.departure_time or "",
            a.itinerary_link or "",
        ])

    # Grand total rows: hotel cost + hotel nights
    grand        = sum((a.hotel_cost or 0) for a in assignments)
    grand_nights = sum((a.stay_nights or 0) for a in assignments)
    ws.append([])
    cost_row = ws.max_row + 1
    ws.cell(row=cost_row, column=9, value="Grand-total hotel cost:").font = Font(bold=True)
    ws.cell(row=cost_row, column=10, value=grand).font = Font(bold=True)
    nights_row = cost_row + 1
    ws.cell(row=nights_row, column=7, value="Grand-total hotel nights:").font = Font(bold=True)
    ws.cell(row=nights_row, column=8, value=grand_nights).font = Font(bold=True)

    widths = [22, 18, 18, 14,   22, 12, 12, 8, 18, 12,   16, 10, 16, 10, 40]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    fname = f"{_slugify(show.code or show.name)}_travel.xlsx"
    return _xlsx_response(wb, fname)


@show_crew_bp.route("/<int:show_id>/crew/travel/print")
def travel_print(show_id):
    """Print-friendly Travel view. Renders the same on-screen page with a
    print stylesheet that hides the sidebar/nav/buttons. Users hit Cmd+P
    (Ctrl+P on Windows) to save as PDF via their browser.

    Replaces the previous WeasyPrint route, which failed on PythonAnywhere
    because Pango/GObject native libs aren't installed there. Browser-side
    printing is portable, reliable, and needs no server deps."""
    show = Show.query.get_or_404(show_id)
    sort_by = (request.args.get("sort") or "check_in").strip().lower()
    if sort_by not in ("check_in", "name", "company", "position"):
        sort_by = "check_in"
    assignments = _travel_assignments_sorted(show_id, sort_by)
    grand_total  = sum((a.hotel_cost or 0) for a in assignments)
    grand_nights = sum((a.stay_nights or 0) for a in assignments)
    return render_template(
        "shows/show_crew_travel.html",
        show=show,
        assignments=assignments,
        grand_total=grand_total,
        grand_nights=grand_nights,
        company_counts=_company_counts(assignments),
        sort_by=sort_by,
        print_mode=True,
    )
