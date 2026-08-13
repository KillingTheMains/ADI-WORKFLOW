"""
Master OSS → XLSX.

Client-facing, so the tests care about document structure as much as data:
the day banding, page breaks landing on day boundaries, and department
identity being consistent.
"""
import datetime as dt
import io


def _show(db):
    from models import (Show, ScheduleDay, ScheduleActivity, SubScheduleEntry,
                        HardCodedEvent)
    show = Show(name="Export Show", code="XL26")
    db.session.add(show); db.session.flush()
    d1 = ScheduleDay(show_id=show.id, date=dt.date(2026, 12, 1),
                     sod="08:00", eod="20:00")
    d2 = ScheduleDay(show_id=show.id, date=dt.date(2026, 12, 2),
                     sod="08:00", eod="20:00")
    db.session.add_all([d1, d2]); db.session.flush()
    db.session.add(ScheduleActivity(day_id=d1.id, time="08:00",
                                    description="LOAD IN", sort_order=10))
    db.session.add(SubScheduleEntry(show_id=show.id, schedule_day_id=d1.id,
                                    type="Dock", time="13:00",
                                    activity="DOCK PUSH", count=3,
                                    duration_hrs=1.5, sort_order=0))
    db.session.add(SubScheduleEntry(show_id=show.id, schedule_day_id=d2.id,
                                    type="House LX", time="09:00",
                                    activity="HOUSE UP", sort_order=0))
    db.session.commit()
    return show


def _workbook(db, show):
    import openpyxl
    from models import SubScheduleEntry, MealService, AgencySetting
    from oss_xlsx import build_workbook
    entries = SubScheduleEntry.query.filter_by(show_id=show.id).all()
    meals = MealService.query.filter_by(show_id=show.id).all()
    return build_workbook(show, entries, meals, agency=AgencySetting.get())


def test_route_returns_a_workbook(app, client, db):
    show = _show(db)
    r = client.get("/shows/%d/oss/master.xlsx" % show.id)
    assert r.status_code == 200
    assert r.mimetype == ("application/vnd.openxmlformats-"
                          "officedocument.spreadsheetml.sheet")
    assert "XL26_Master_Schedule_" in r.headers["Content-Disposition"]
    assert r.data[:2] == b"PK"          # a real zip-backed xlsx


def test_expected_sheets_exist(app, db):
    show = _show(db)
    wb = _workbook(db, show)
    assert wb.sheetnames[0] == "Cover"
    assert wb.sheetnames[1] == "Master Schedule"
    assert wb.sheetnames[-1] == "Summary"
    assert "Dock" in wb.sheetnames


def test_department_identity_is_normalised(app, db):
    """'House LX' is the stored TYPE; users see 'House Lights'. A department
    must never appear twice under two names, or split into two sheets."""
    show = _show(db)
    wb = _workbook(db, show)
    assert "House Lights" in wb.sheetnames
    assert "House LX" not in wb.sheetnames
    # Department moved to column 3 behind the kind-code column (§09).
    depts = {wb["Master Schedule"].cell(row=r, column=3).value
             for r in range(3, wb["Master Schedule"].max_row + 1)}
    assert "House LX" not in depts


def test_master_is_banded_by_day_and_breaks_on_day_boundaries(app, db):
    show = _show(db)
    ws = _workbook(db, show)["Master Schedule"]

    # Day banners use the house date format — "Tue 1 Dec 2026" (Larry, 9 Aug).
    banners = [(r, ws.cell(row=r, column=1).value)
               for r in range(3, ws.max_row + 1)
               if str(ws.cell(row=r, column=1).value or "").startswith(
                   ("Mon ", "Tue ", "Wed ", "Thu ", "Fri ", "Sat ", "Sun "))]
    assert len(banners) == 2, f"expected one banner per day, got {banners}"

    # A page break immediately before every day after the first, so a day
    # never starts halfway down a printed page.
    breaks = [b.id for b in ws.row_breaks.brk]
    assert breaks == [banners[1][0] - 1], f"breaks {breaks} vs banners {banners}"


def test_master_has_no_autofilter_over_merged_banners(app, db):
    """Excel handles a filter across merged cells badly — filtering lives on
    the flat department sheets instead."""
    show = _show(db)
    wb = _workbook(db, show)
    assert wb["Master Schedule"].auto_filter.ref is None
    assert wb["Dock"].auto_filter.ref is not None


def test_print_setup_is_configured_for_pdf_export(app, db):
    show = _show(db)
    ws = _workbook(db, show)["Master Schedule"]
    assert ws.print_title_rows == "$1:$2"      # header repeats on every page
    assert ws.page_setup.fitToWidth == 1
    assert ws.freeze_panes == "A3"


def test_rows_are_chronological_within_each_day(app, db):
    from time_utils import parse_minutes
    show = _show(db)
    ws = _workbook(db, show)["Master Schedule"]
    run, runs = [], []
    # Column indices shifted +1 when the kind-code column 'K' became
    # column 1 of the Master sheet (Interface Spec §09). Behaviour is
    # unchanged — only which column carries it moved.
    for r in range(3, ws.max_row + 1):
        # A day banner is merged from column 1, so its Department cell is
        # empty; an item row always has one. Time is column 2 now.
        if ws.cell(row=r, column=3).value is None:
            if run: runs.append(run)
            run = []
        else:
            run.append(parse_minutes(ws.cell(row=r, column=2).value))
    if run: runs.append(run)
    for run in runs:
        assert run == sorted(run), f"day is out of clock order: {run}"


def test_detail_column_folds_count_and_duration(app, db):
    show = _show(db)
    ws = _workbook(db, show)["Master Schedule"]
    # Detail moved 4 -> 5 behind the kind-code column (§09).
    details = [ws.cell(row=r, column=5).value for r in range(3, ws.max_row + 1)]
    assert any(d and "×3" in d and "1.5 hr" in d for d in details), details


def test_summary_counts_with_formulas_not_baked_numbers(app, db):
    """Figures stay true if anyone filters or edits the delivered sheet."""
    show = _show(db)
    ws = _workbook(db, show)["Summary"]
    formulas = [ws.cell(row=r, column=2).value for r in range(4, ws.max_row + 1)]
    assert any(str(f).startswith("=COUNTIF('Master Schedule'!") for f in formulas)
    assert any(str(f).startswith("=SUM(") for f in formulas)


def test_the_kind_code_column_leads_the_master_sheet(app, db):
    """Interface Spec §09: the two-letter code is the ONE channel that carries
    row kind through greyscale, a photocopier and a fax.

    On screen a kind is told by a chip, a rail pattern AND a silhouette. Paper
    keeps only the chip, so if this column is wrong or missing there is nothing
    left — which is exactly why it is worth a test of its own rather than
    riding along on the layout assertions above.
    """
    import brand
    show = _show(db)
    ws = _workbook(db, show)["Master Schedule"]

    assert ws.cell(row=2, column=1).value == "K", "the code column leads"

    codes = [ws.cell(row=r, column=1).value
             for r in range(3, ws.max_row + 1)
             if ws.cell(row=r, column=3).value]      # skip day banners
    assert codes, "every item row carries a code"

    # Department codes are legal codes on this sheet too — an OSS entry
    # prints DK rather than AC, via brand.row_code(), so paper says what the
    # screen says.
    known = set(brand.KIND_CODE.values()) | set(brand.DEPT_CODE.values())
    assert all(c in known for c in codes), \
        f"unknown code printed: {sorted(set(codes) - known)}"

    # The codes must agree with brand.KIND_CODE rather than being spelled out
    # again here — oss_pdf and oss_xlsx printing DIFFERENT letters for the same
    # row would be worse than printing none.
    assert brand.KIND_CODE["break"] == "BR" and brand.KIND_CODE["crew"] == "CC"
