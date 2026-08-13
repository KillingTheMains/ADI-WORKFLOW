"""
Master OSS → XLSX.

Client-facing, so it is built as a document rather than a data dump: a cover
sheet, the master timeline banded by day, a sheet per department, and a
summary. Reads from oss_export.build_master_items so it can never disagree
with what the Master tab shows.

Print setup is deliberate — page breaks land on day boundaries and the header
row repeats — so Excel's own Save-as-PDF gives a usable document even though
the native PDF export is the better route.
"""
import os
from datetime import datetime

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
try:                       # openpyxl 3.x
    from openpyxl.worksheet.pagebreak import Break as RowBreak
except ImportError:        # older releases called it Brk
    from openpyxl.worksheet.pagebreak import Brk as RowBreak
from openpyxl.worksheet.properties import PageSetupProperties

import brand
from oss_export import (build_master_items, count_label, department_style,
                        group_by_day, group_by_department, master_label,
                        time_range_text)

FONT = brand.FONT_FALLBACK_XLSX
WHITE = "FFFFFF"
RULE = "D9DEE7"
BAND = "F4F6F9"
DEFAULT_NAVY = brand.as_openpyxl(brand.PRIMARY)


# The kind code leads every timeline-shaped sheet, exactly as it leads the
# PDF's tables and the day page's rows. Interface Spec §07/§09.
MASTER_HEADERS = ["K", "Time", "Department", "Item", "Detail", "Notes"]
KIND_FILLS = {k: PatternFill("solid", fgColor=v.lstrip("#"))
              for k, v in brand.KIND_FILL.items()}
DEPT_HEADERS = ["Day", "Date", "Time", "Item", "Detail", "Count", "Hrs", "Notes"]

_thin = Side(style="thin", color=RULE)
BORDER = Border(bottom=_thin)


def _navy(agency):
    """Brand colour as a bare RRGGBB, which is what openpyxl wants."""
    raw = (getattr(agency, "primary_hex", None) or "").lstrip("#")
    return raw.upper() if len(raw) == 6 else DEFAULT_NAVY


def _detail(item):
    """Count and duration folded into one readable cell.

    The master's rows have very different shapes — a Crew row is a list of
    names, an F&B row is a service plus location, a Dock row has a count and a
    duration. Separate columns for each would leave the client-facing sheet
    mostly empty, so the operational fields fold in here and the department
    sheets carry them properly.
    """
    bits = []
    if item.get("count") is not None:
        bits.append(count_label(item["dept"], item["count"])
                    or f"×{item['count']}")
    if item.get("duration_hrs") is not None:
        bits.append(f"{item['duration_hrs']:g} hr")
    return " · ".join(bits)


def _display_time(value):
    """24-hour on generated documents (Larry, 2026-08-09). The on-screen app
    still shows 12-hour — that is deliberate, screens and client documents
    have different readers."""
    return brand.fmt_time(value)


def _day_label(day):
    if day is None or not day.date:
        return "Unscheduled"
    label = brand.fmt_date(day.date, "full")
    if getattr(day, "phase_labels", None):
        labels = day.phase_labels
        if labels:
            label += "  ·  " + (", ".join(labels) if isinstance(labels, (list, tuple))
                                else str(labels))
    return label


def _fit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _print_setup(ws, landscape=False, repeat="1:1"):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = repeat
    ws.print_options.horizontalCentered = True


def _cover(wb, show, agency, master_items, logo_file):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    navy = _navy(agency)
    _fit(ws, [3, 26, 26, 26, 3])

    # Navy panel across the top — the reversed logo sits on it natively.
    for row in range(1, 11):
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=navy)
        ws.row_dimensions[row].height = 20

    if logo_file and os.path.exists(logo_file):
        try:
            img = XLImage(logo_file)
            scale = min(280 / max(img.width, 1), 90 / max(img.height, 1), 1)
            img.width, img.height = int(img.width * scale), int(img.height * scale)
            ws.add_image(img, "B2")
        except Exception:
            pass                                   # never fail an export over art

    title = ws.cell(row=8, column=2, value=show.name or "")
    title.font = Font(name=FONT, size=22, bold=True, color=WHITE)
    sub = ws.cell(row=9, column=2,
                  value=" · ".join(filter(None, [show.code, "Master Schedule"])))
    sub.font = Font(name=FONT, size=11, color="AEBBCC")


    def _fact(row, label, value):
        k = ws.cell(row=row, column=2, value=label)
        k.font = Font(name=FONT, size=9, bold=True, color="6B7280")
        v = ws.cell(row=row, column=3, value=value or "—")
        v.font = Font(name=FONT, size=11)
        v.alignment = Alignment(horizontal="left")

    days = [d for d, _ in group_by_day(master_items) if d and d.date]
    span = ""
    if days:
        first, last = days[0].date, days[-1].date
        span = (brand.fmt_date(first, "full") if first == last else
                f"{brand.fmt_date(first)} – {brand.fmt_date(last, 'full')}")

    _fact(13, "CLIENT", show.client.name if show.client else None)
    _fact(14, "VENUE", show.venue.name if show.venue else None)
    _fact(15, "ROOM", show.room_name)
    _fact(16, "DATES", span)
    _fact(17, "SCHEDULE DAYS", str(len(days)) if days else "0")
    _fact(18, "TOTAL ITEMS", str(len(master_items)))
    _fact(19, "PREPARED BY", getattr(agency, "name", None) or "ADI Productions")
    _fact(20, "GENERATED", datetime.now().strftime("%-d %b %Y, %H:%M"))  # 24h

    key = ws.cell(row=23, column=2, value="DEPARTMENT KEY")
    key.font = Font(name=FONT, size=9, bold=True, color="6B7280")
    row = 24
    for dept, _rows in group_by_department(master_items):
        style = department_style(dept)
        chip = ws.cell(row=row, column=2, value=style["short"] or dept[:5])
        chip.fill = PatternFill("solid", fgColor=style["hex"])
        chip.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        chip.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=dept).font = Font(name=FONT, size=10)
        row += 1
    _print_setup(ws)
    return ws


def _master(wb, show, agency, master_items):
    ws = wb.create_sheet("Master Schedule")
    ws.sheet_view.showGridLines = False
    navy = _navy(agency)
    _fit(ws, [4, 11, 15, 42, 20, 44])

    head = ws.cell(row=1, column=1,
                   value=f"{show.code or ''}  {show.name or ''}  —  Master Schedule".strip())
    head.font = Font(name=FONT, size=13, bold=True, color=navy)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.row_dimensions[1].height = 22

    for col, label in enumerate(MASTER_HEADERS, start=1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=navy)
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    row = 3
    day_break_rows = []
    for day, items in group_by_day(master_items):
        if row > 3:
            day_break_rows.append(row - 1)     # start each day on a fresh page
        banner = ws.cell(row=row, column=1, value=_day_label(day))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        banner.font = Font(name=FONT, size=11, bold=True, color=WHITE)
        banner.fill = PatternFill("solid", fgColor=navy)
        banner.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1

        for n, item in enumerate(items):
            style = department_style(item["dept"])
            kind = item.get("kind") or "act"
            values = [brand.KIND_CODE.get(kind, ""),
                      time_range_text(item, _display_time), item["dept"],
                      master_label(item), _detail(item), item["notes"]]
            # The band alternated by POSITION, so it carried nothing and
            # printed as noise. It now carries KIND, from the same three-tier
            # map the PDF uses.
            kind_fill = KIND_FILLS.get(kind)
            for col, value in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=value)
                c.font = Font(name=FONT, size=10)
                c.border = BORDER
                c.alignment = Alignment(vertical="top",
                                        wrap_text=(col in (4, 6)))
                if kind_fill:
                    c.fill = kind_fill
            ws.cell(row=row, column=1).font = Font(name=FONT, size=9, bold=True)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center",
                                                             vertical="top")
            ws.cell(row=row, column=2).font = Font(name=FONT, size=10, bold=True)
            # Department cell carries the accent, so a printed page still
            # groups visually even without the chip colours.
            ws.cell(row=row, column=3).font = Font(name=FONT, size=10,
                                                   bold=True, color=style["hex"])
            row += 1

            # Note 5 — Larry wants a tall list, one name per row, matching the
            # show book's call sheet. The headcount above stays: it is how the
            # count is read at a glance, and it is load-bearing in the workflow.
            for who in item.get("crew_names") or []:
                c = ws.cell(row=row, column=4, value=who)
                c.font = Font(name=FONT, size=10)
                c.alignment = Alignment(vertical="top", indent=2)
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = BORDER
                    if kind_fill:
                        cell.fill = kind_fill
                row += 1

            # LOCAL LABOUR, split out of the name list. A line reading
            # "14 × Rigger" used to be written into the same column with the
            # same font and an empty code cell — one filter on this sheet
            # could not tell fourteen people from one. It now carries LL and
            # the local-labour fill, and its headcount goes in the detail
            # column so the number is filterable.
            local_fill = KIND_FILLS.get("local")
            for line in item.get("local_lines") or []:
                ws.cell(row=row, column=1,
                        value=brand.KIND_CODE.get("local", "")).font = Font(
                            name=FONT, size=9, bold=True)
                ws.cell(row=row, column=1).alignment = Alignment(
                    horizontal="center", vertical="top")
                c = ws.cell(row=row, column=4, value=line.get("label"))
                c.font = Font(name=FONT, size=10)
                c.alignment = Alignment(vertical="top", indent=2)
                d = ws.cell(row=row, column=5,
                            value=count_label("Crew", line.get("qty")))
                d.font = Font(name=FONT, size=10)
                d.alignment = Alignment(vertical="top")
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = BORDER
                    if local_fill:
                        cell.fill = local_fill
                row += 1

    last = row - 1
    # Deliberately NO autofilter here: the day banners are merged across the
    # row and Excel handles a filter over merged cells badly. The flat
    # department sheets carry the filtering instead.
    for r in day_break_rows:
        ws.row_breaks.append(RowBreak(id=r, man=True))
    _print_setup(ws, repeat="1:2")
    return ws, last


def _safe_sheet_name(name, used):
    """Excel tab names: 31 chars, and none of : \\ / ? * [ ]"""
    clean = "".join("-" if ch in ':\\/?*[]' else ch for ch in name)[:31] or "Sheet"
    candidate, n = clean, 2
    while candidate.lower() in used:
        suffix = f" {n}"
        candidate = clean[:31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def _department_sheets(wb, show, agency, master_items):
    """One tab per department, carrying the full field set — the operational
    view, where the Master sheet stays deliberately readable."""
    used = {"cover", "master schedule", "summary"}
    made = []
    for dept, items in group_by_department(master_items):
        style = department_style(dept)
        ws = wb.create_sheet(_safe_sheet_name(dept, used))
        ws.sheet_view.showGridLines = False
        _fit(ws, [18, 12, 11, 40, 20, 8, 7, 40])

        title = ws.cell(row=1, column=1, value=f"{dept} — {show.name or ''}".strip())
        title.font = Font(name=FONT, size=12, bold=True, color=style["hex"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        for col, label in enumerate(DEPT_HEADERS, start=1):
            c = ws.cell(row=2, column=col, value=label)
            c.font = Font(name=FONT, size=9, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=style["hex"])
        ws.freeze_panes = "A3"

        row = 3
        for n, item in enumerate(items):
            day = item["day"]
            # The Crew sheet is where the names live — the Master shows a count.
            # Local labour is joined in after the people rather than dropped:
            # this is a flat, filterable sheet with one row per item and no
            # room for a second row per line, so the alternative would be a
            # Crew tab that silently omits every local-labour call.
            names = list(item.get("crew_names") or [])
            names += [l.get("label") for l in item.get("local_lines") or []]
            label = (", ".join(names) if names else item["activity"])
            values = [
                brand.fmt_date(day.date) if day and day.date else "Unscheduled",
                day.date.isoformat() if day and day.date else "",
                time_range_text(item, _display_time),
                label,
                count_label(item["dept"], item.get("count")),
                item.get("count"),
                item.get("duration_hrs"),
                item["notes"],
            ]
            for col, value in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=value)
                c.font = Font(name=FONT, size=10)
                c.border = BORDER
                c.alignment = Alignment(vertical="top",
                                        wrap_text=(col in (4, 8)))
                if n % 2:
                    c.fill = PatternFill("solid", fgColor=BAND)
            row += 1

        if row > 3:
            ws.auto_filter.ref = f"A2:H{row - 1}"
        _print_setup(ws, landscape=True, repeat="1:2")
        made.append((ws.title, dept, len(items)))
    return made


def _summary(wb, show, agency, master_items, master_last_row):
    """Counts by department and by day.

    COUNTIF against the Master sheet rather than numbers baked in from Python,
    so the figures stay true if anyone filters or edits the export.
    """
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    navy = _navy(agency)
    _fit(ws, [30, 12, 6, 30, 12])

    t = ws.cell(row=1, column=1, value="Schedule summary")
    t.font = Font(name=FONT, size=12, bold=True, color=navy)

    for col, label in ((1, "Department"), (2, "Items")):
        c = ws.cell(row=3, column=col, value=label)
        c.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=navy)

    row = 4
    rng = f"'Master Schedule'!$B$3:$B${max(master_last_row, 3)}"
    for dept, _items in group_by_department(master_items):
        ws.cell(row=row, column=1, value=dept).font = Font(name=FONT, size=10)
        c = ws.cell(row=row, column=2, value=f'=COUNTIF({rng},A{row})')
        c.font = Font(name=FONT, size=10)
        row += 1
    total = ws.cell(row=row, column=2, value=f"=SUM(B4:B{row - 1})")
    total.font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=row, column=1, value="Total").font = Font(name=FONT, size=10, bold=True)

    for col, label in ((4, "Day"), (5, "Items")):
        c = ws.cell(row=3, column=col, value=label)
        c.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=navy)
    row = 4
    for day, items in group_by_day(master_items):
        ws.cell(row=row, column=4,
                value=_day_label(day)).font = Font(name=FONT, size=10)
        ws.cell(row=row, column=5,
                value=len(items)).font = Font(name=FONT, size=10)
        row += 1
    _print_setup(ws)
    return ws


def build_workbook(show, entries, meal_services, agency=None, logo_file=None):
    """The finished workbook. Caller supplies the already-queried collections
    so this stays a pure formatting layer over oss_export."""
    master_items, _hardcoded = build_master_items(show, entries, meal_services)

    wb = openpyxl.Workbook()
    wb.properties.title = f"{show.code or ''} {show.name or ''} — Master Schedule".strip()
    wb.properties.creator = getattr(agency, "name", None) or "ADI Productions"

    _cover(wb, show, agency, master_items, logo_file)
    _master_ws, last_row = _master(wb, show, agency, master_items)
    _department_sheets(wb, show, agency, master_items)
    _summary(wb, show, agency, master_items, last_row)
    wb.active = 1                       # open on the schedule, not the cover
    return wb
